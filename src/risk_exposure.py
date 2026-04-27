import csv
import os
import re
import shutil
import time
import warnings
from datetime import datetime

import matplotlib.pyplot as plt
import nltk
import pandas as pd
import requests
import seaborn as sns
from bs4 import BeautifulSoup, XMLParsedAsHTMLWarning
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer

warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
from nltk.tokenize import word_tokenize
from sec_edgar_downloader import Downloader

nltk.download('punkt')
nltk.download('punkt_tab')
nltk.download('stopwords')
nltk.download('wordnet')
nltk.download('omw-1.4')

# 1. READ UNIQUE CIKS from file
print("Reading CIKs from C:\\Users\\Admin\\Downloads\\New folder\\ciks.txt...")
unique_ciks = []
try:
    with open(r"C:\Users\Admin\Downloads\New folder\ciks.txt", "r") as f:
        unique_ciks = sorted(list(set([line.strip() for line in f.readlines() if line.strip()])))
    print(f"Found {len(unique_ciks)} unique CIKs.")
except Exception as e:
    print(f"Error reading ciks.txt: {e}")
    unique_ciks = ["0000320193"] # Fallback

risk_keywords = [
    "Data center", "Server farm", "Colocation", "Hyperscaler", "Edge computing",
    "Cloud infrastructure", "Data warehousing", "Cloud computing", "GPU cluster",
    "Infrastructure as a Service", "Low latency", "High bandwidth", "Redundancy",
    "Scalability", "Power usage effectiveness", "Uptime", "Data sovereignty",
    "Disaster recovery", "Cybersecurity infrastructure", "Network infrastructure",
    "Artificial Intelligence", "Machine learning", "Big data analytics",
    "Digital transformation", "Internet of Things", "Omnichannel strategy",
    "High-frequency trading", "Streaming infrastructure", "Content Delivery Network",
    "FinTech platform", "ICT investment", "Digital asset", 
    "Capital expenditure in technology", "Digital infrastructure",
    "Real estate technology investment", "Tech-intensive assets",
    "Infrastructure investment", "Technological innovation",
    "IT capital allocation", "Facility modernization"
]

patterns = {f"{k.replace(' ', '_').lower()}_count": rf"\b{re.escape(k)}s?\b" for k in risk_keywords}


# 2. TEXT PREPROCESSING FUNCTION
def preprocess_text(soup):
    for table in soup.find_all("table"):
        table.decompose()
    text = soup.get_text(separator=' ').lower()

    contractions = {
        "don't": "do not", "can't": "cannot", "it's": "it is", "isn't": "is not",
        "aren't": "are not", "won't": "will not", "shouldn't": "should not",
        "i'm": "i am", "you're": "you are", "we're": "we are", "they're": "they are"
    }
    for contraction, expanded in contractions.items():
        text = text.replace(contraction, expanded)

    text = re.sub(r'https?://\S+|www\.\S+', '', text)
    text = re.sub(r'[^a-z\s]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()

    tokens = word_tokenize(text)
    stop_words = set(stopwords.words('english'))
    lemmatizer = WordNetLemmatizer()
    
    tokens = [lemmatizer.lemmatize(w) for w in tokens if w not in stop_words and len(w) > 2]
    return " ".join(tokens)


# 3. ANALYSIS PIPELINE
print("Fetching company mapping from SEC...")
headers = {"User-Agent": "harshita@emergeflow.com"}
try:
    mapping_data = requests.get("https://www.sec.gov/files/company_tickers.json", headers=headers).json()
    cik_lookup = {str(entry['cik_str']): {'name': entry['title'], 'ticker': entry['ticker']} for entry in mapping_data.values()}
except:
    print("Failed to fetch company tickers.")
    cik_lookup = {}

dl = Downloader("Emergeflow-Technologies", "harshita@emergeflow.com")
csv_filename = r"C:\Users\Admin\Downloads\New folder\comprehensive_risk_analysis.csv"

# --- RESUME LOGIC ---
processed_records = set()
if os.path.exists(csv_filename):
    try:
        df_existing = pd.read_csv(csv_filename, usecols=['cik', 'Year'], dtype={'cik': str})
        for _, row in df_existing.iterrows():
            processed_records.add((str(row['cik']).lstrip('0'), int(row['Year'])))
        print(f"Found {len(processed_records)} already processed records in {csv_filename}. These will be skipped.")
    except Exception as e:
        print(f"Could not read existing CSV for resumption: {e}")

output_columns = ["cik", "ticker", "coname", "Year", "word_count"] + list(patterns.keys()) + [k.replace('_count', '_norm') for k in patterns.keys()]

# --- AUTO-TIMING SETUP ---
log_filename = r"C:\Users\Admin\Downloads\New folder\processing_log.csv"
log_columns = ["timestamp", "year", "batch_start", "batch_end",
               "ciks_in_batch", "ciks_processed", "batch_time_seconds",
               "cumulative_ciks", "total_ciks_target",
               "estimated_remaining_hours"]
if not os.path.exists(log_filename):
    with open(log_filename, "w", newline="") as lf:
        csv.writer(lf).writerow(log_columns)

total_ciks_target = len(unique_ciks) * 27  # 27 years
cumulative_ciks_done = len(processed_records)
batch_times = []  # running list of batch durations for ETA
script_start_time = time.time()

for year in range(2000, 2027):
    print(f"\n{'='*40}")
    print(f"PROCESSING YEAR: {year}")
    print(f"{'='*40}")
    
    year_start_time = time.time()
    year_ciks_done = 0

    for i in range(0, len(unique_ciks), 100):
        batch_ciks = unique_ciks[i:i+100]
        
        # Fast batch skip Check
        all_skipped = True
        for cik in batch_ciks:
            if (str(cik).lstrip('0'), year) not in processed_records:
                all_skipped = False
                break
                
        if all_skipped:
            print(f"--- Skipping batch {min(i+100, len(unique_ciks))}/{len(unique_ciks)} for year {year} (all CIKs already processed) ---")
            continue
            
        print(f"\n--- Processing batch {min(i+100, len(unique_ciks))}/{len(unique_ciks)} for year {year} ---")
        batch_start_time = time.time()
        batch_cik_count = 0
        
        for cik in batch_ciks:
            raw_cik_str = str(cik).lstrip('0')
            if (raw_cik_str, year) in processed_records:
                continue
                
            # Ensure completely clean state before downloading to prevent leftover files polluting the data
            base_dir = "sec-edgar-filings"
            if os.path.exists(base_dir):
                try:
                    shutil.rmtree(base_dir)
                except Exception:
                    pass
                    
            try:
                print(f"  Downloading CIK {cik}...")
                dl.get("10-K", cik, after=f"{year}-01-01", before=f"{year+1}-01-01", download_details=False)
                time.sleep(0.15)
            except Exception as e:
                print(f"  Failed to download for {cik}: {e}")
        
            # Preprocess immediately
            base_dir = "sec-edgar-filings"
            results = []
            if os.path.exists(base_dir):
                for root, dirs, files in os.walk(base_dir):
                    for file in files:
                        if file.endswith((".html", ".txt")):
                            file_path = os.path.join(root, file)
                            path_parts = file_path.split(os.sep)
                            
                            raw_cik = ""
                            for part in path_parts:
                                if len(part) == 10 and part.isdigit():
                                    raw_cik = part.lstrip('0')
                                    break
                            if not raw_cik: continue
                            
                            company_info = cik_lookup.get(raw_cik, {'name': 'Unknown', 'ticker': 'N/A'})
                            
                            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                file_content = f.read()
                                if not file_content.strip(): continue
                                
                                processed_text = preprocess_text(BeautifulSoup(file_content, 'lxml'))
                                word_count = len(processed_text.split())
                                if word_count == 0: continue
            
                                row_data = {
                                    "cik": raw_cik, "ticker": company_info['ticker'], "coname": company_info['name'],
                                    "Year": year, "word_count": word_count
                                }
                                
                                for label, regex in patterns.items():
                                    count = len(re.findall(regex, processed_text, re.IGNORECASE))
                                    row_data[label] = count
                                    norm_name = label.replace("_count", "_norm")
                                    row_data[norm_name] = (count / word_count) * 10000
                                
                                results.append(row_data)

                # Cleanup immediately so it doesn't bleed to the next CIK
                try:
                    shutil.rmtree(base_dir)
                except Exception as e:
                    pass
            
            if not results:
                # Need to add placeholder to bypass this in the future
                results.append({
                    "cik": raw_cik_str,
                    "ticker": cik_lookup.get(raw_cik_str, {}).get("ticker", "N/A"),
                    "coname": cik_lookup.get(raw_cik_str, {}).get("name", "Unknown"),
                    "Year": year,
                    "word_count": -1 # Special negative value to denote empty result
                })
                
            df_batch = pd.DataFrame(results)
            for col in output_columns:
                if col not in df_batch.columns:
                    df_batch[col] = 0
            df_batch = df_batch[output_columns]
            
            write_header = not os.path.exists(csv_filename)
            df_batch.to_csv(csv_filename, mode='a', index=False, header=write_header)
            
            processed_records.add((raw_cik_str, year))
            batch_cik_count += 1
            cumulative_ciks_done += 1
            year_ciks_done += 1
            
        # --- LOG BATCH TIMING ---
        batch_elapsed = time.time() - batch_start_time
        batch_times.append(batch_elapsed)
        avg_batch_time = sum(batch_times) / len(batch_times)
        remaining_ciks = total_ciks_target - cumulative_ciks_done
        remaining_batches = remaining_ciks / max(len(batch_ciks), 1)
        est_remaining_hours = (remaining_batches * avg_batch_time) / 3600

        log_row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            year, i, min(i + 100, len(unique_ciks)),
            len(batch_ciks), batch_cik_count,
            round(batch_elapsed, 2),
            cumulative_ciks_done, total_ciks_target,
            round(est_remaining_hours, 2)
        ]
        with open(log_filename, "a", newline="") as lf:
            csv.writer(lf).writerow(log_row)

        print(f"-> COMPLETED: Batch {min(i+100, len(unique_ciks))} in year {year}. "
              f"Time: {batch_elapsed:.1f}s | Cumulative: {cumulative_ciks_done}/{total_ciks_target} "
              f"| ETA: {est_remaining_hours:.1f}h")

    # --- YEAR SUMMARY ---
    year_elapsed = time.time() - year_start_time
    print(f"\n*** YEAR {year} COMPLETE -- {year_ciks_done} CIKs in {year_elapsed:.0f}s ({year_elapsed/60:.1f}min) ***")

print("\nAll data processing is finished.")

if os.path.exists(csv_filename):
    final_df = pd.read_csv(csv_filename)
    final_df = final_df[final_df['word_count'] > 0]  # Filter out missing-data placeholders
    norm_cols = [c for c in final_df.columns if "_norm" in c]
    
    if not final_df.empty:
        with open(r"C:\Users\Admin\Downloads\New folder\top_10_risk_companies.txt", "w") as f:
            for norm_col in norm_cols:
                f.write(f"\n--- TOP 10 FOR: {norm_col.upper()} ---\n")
                top_10 = final_df.sort_values(by=norm_col, ascending=False).head(10)
                f.write(top_10[['coname', 'ticker', norm_col]].to_string(index=False) + "\n")
        
        print("Generating plots...")
        for norm_col in norm_cols:
            plt.figure(figsize=(10, 4))
            yearly_trend = final_df.groupby('Year')[norm_col].mean()
            risk_name = norm_col.replace('_norm', '').replace('_', ' ').title()
            
            plt.plot(yearly_trend.index, yearly_trend.values, marker='o', color='teal')
            plt.title(f"Trend for Risk Exposure: {risk_name} Over the Years")
            plt.xlabel("Year")
            plt.ylabel("Avg Mentions per 10k words")
            plt.grid(True, alpha=0.3)
            plt.xticks(yearly_trend.index, rotation=45)
            plt.tight_layout()
            plt.savefig(rf"C:\Users\Admin\Downloads\New folder\{norm_col}_trend.png")
            plt.close()

        plt.figure(figsize=(15, 12))
        corr = final_df[norm_cols].corr()
        sns.heatmap(corr, annot=False, cmap='coolwarm', linewidths=0.1)
        plt.title("Correlation Matrix of Risk Exposure Keywords")
        plt.tight_layout()
        plt.savefig(r"C:\Users\Admin\Downloads\New folder\correlation_matrix.png")
        plt.close()

        print("!!! ANALYSIS AND PLOTTING COMPLETE !!! Check 'top_10_risk_companies.txt' for leaderboards.")
else:
    print("No data was collected during the processing phase.")

# ============================================================
# AUTO-UPDATE EXCEL TASK TRACKER
# ============================================================
def update_excel_tracker():
    """Append a daily-progress row & update task statuses in the xlsx tracker."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[WARN] openpyxl not installed -- skipping Excel update.")
        return

    xlsx_path = r"C:\Users\Admin\Downloads\New folder\SEC_Risk_Exposure_Task_Tracker.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"[WARN] Excel tracker not found at {xlsx_path} -- skipping.")
        return

    wb = load_workbook(xlsx_path)

    # --- 1. Append row to "Daily Progress" sheet ---
    if "Daily Progress" in wb.sheetnames:
        ws_daily = wb["Daily Progress"]
        total_batches_done = len(batch_times)
        avg_time = sum(batch_times) / max(len(batch_times), 1)
        remaining = total_ciks_target - cumulative_ciks_done
        remaining_b = remaining / 100
        eta_hrs = (remaining_b * avg_time) / 3600 if batch_times else 0
        script_elapsed = time.time() - script_start_time
        notes = (f"Run duration: {script_elapsed/60:.1f}min | "
                 f"Processed {cumulative_ciks_done}/{total_ciks_target} CIK-years")
        ws_daily.append([
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            total_batches_done,
            cumulative_ciks_done,
            round(avg_time, 2),
            round(eta_hrs, 2),
            notes
        ])
        print(f"[INFO] Daily progress row appended to Excel.")

    # --- 2. Update task statuses in "Task Breakdown" sheet ---
    if "Task Breakdown" in wb.sheetnames:
        ws_tasks = wb["Task Breakdown"]
        pct_done = (cumulative_ciks_done / max(total_ciks_target, 1)) * 100
        for row in ws_tasks.iter_rows(min_row=2, max_col=5):
            task_cell = row[1]  # Column B = Tasks
            status_cell = row[4]  # Column E = Status
            if task_cell.value and "1.3" in str(task_cell.value):
                if pct_done >= 100:
                    status_cell.value = "Complete"
                else:
                    status_cell.value = f"In Progress ({pct_done:.1f}%)"
            if task_cell.value and "1.7" in str(task_cell.value):
                status_cell.value = "Complete"
            if task_cell.value and "6.3" in str(task_cell.value):
                status_cell.value = "Complete"

    wb.save(xlsx_path)
    print(f"[INFO] Excel tracker updated: {xlsx_path}")

update_excel_tracker()

if __name__ == "__main__":
    print("SCRIPT EXECUTION FINISHED")